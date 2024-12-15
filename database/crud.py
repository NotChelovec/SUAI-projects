import datetime

from sqlalchemy import exists, select, delete, Integer, and_
from sqlalchemy import func

from database import crud
from database import database as db
from model.academic_performance import (
    AcademicPerformance,
    MissedInfo,
    FullAcademicPerformance,
)
from model.admin import Admin
from model.group import Group
from model.student import Student
from model.discipline import Discipline
from model.missed_class import MissedClass
from model.assigned_discipline import association_table
from model.team import Team

import openpyxl
import os
from dotenv import load_dotenv
from model.student import Student

from typing import Dict, List

load_dotenv()
PATH_TO_TEAMS = os.getenv('PATH_TO_TEAMS')
PATH_TO_CONFIG_DATA = os.getenv('PATH_TO_CONFIG_DATA')

def is_admin(user_telegram_id: int) -> bool:
    with db.Session() as session:
        admin = session.get(Admin, user_telegram_id)
        return admin is not None


def get_students(group_id: int) -> list[Student]:
    with db.Session() as session:
        smt = select(Student).where(Student.group_id == group_id)
        return session.scalars(smt).all()


def get_groups() -> list[Group]:
    with db.Session() as session:
        smt = select(Group)
        return session.scalars(smt).all()


def get_disciplines() -> list[Discipline]:
    with db.Session() as session:
        smt = select(Discipline)
        return session.scalars(smt).all()


def get_discipline(discipline_id: int) -> Discipline:
    with db.Session() as session:
        return session.get(Discipline, discipline_id)


def get_group(group_id: int) -> Group:
    with db.Session() as session:
        return session.get(Group, group_id)


def get_assigned_discipline(group_id: int) -> list[Discipline]:
    with db.Session() as session:
        group = session.get(Group, group_id)
        return group.disciplines


def get_assigned_group(discipline_id: int) -> list[Group]:
    with db.Session() as session:
        discipline = session.get(Discipline, discipline_id)
        return discipline.groups


def set_missed_students(
    students_id: list[int],
    group_id: int,
    discipline_id: int,
) -> None:
    session = db.Session()
    group = session.get(Group, group_id)
    current_date = datetime.date.today()
    for student in group.students:
        is_missed = student.id in students_id
        student.missed.append(
            MissedClass(
                discipline_id=discipline_id,
                date=current_date,
                is_missed=is_missed,
            )
        )
    session.commit()
    session.close()


def set_all_missed_students(
    group_id: int,
    discipline_id: int,
    is_missed=True,
) -> None:
    session = db.Session()
    group = session.get(Group, group_id)
    current_date = datetime.date.today()
    for student in group.students:
        student.missed.append(
            MissedClass(
                discipline_id=discipline_id,
                date=current_date,
                is_missed=is_missed,
            )
        )
    session.commit()
    session.close()


def rename_student(student_id: int, new_fullname: str) -> None:
    with db.Session() as session:
        student = session.get(Student, student_id)
        student.full_name = new_fullname
        session.commit()


def remove_student(student_id: int) -> None:
    with db.Session() as session:
        smt = delete(Student).where(Student.id == student_id)
        session.execute(smt)
        session.commit()


def remove_group(group_id: int) -> None:
    session = db.Session()
    smt = delete(Group).where(Group.id == group_id)
    session.execute(smt)
    session.commit()
    session.close()


def add_student(group_id: int, full_name: str) -> None:
    with db.Session() as session:
        group = session.get(Group, group_id)
        group.students.append(Student(full_name=full_name))
        session.commit()


def check_discipline(discipline_name: str) -> bool:
    with db.Session() as session:
        smt = select(Discipline).where(
            Discipline.name == discipline_name,
        )
        return session.scalar(smt) is not None


def add_discipline(discipline_name: str) -> None:
    with db.Session() as session:
        discipline = Discipline(name=discipline_name)
        session.add(discipline)
        session.commit()


def check_group(group_name: str) -> bool:
    with db.Session() as session:
        smt = select(Group).where(Group.name == group_name)
        return session.scalar(smt) is not None


def add_group(group_name: str) -> None:
    with db.Session() as session:
        group = Group(name=group_name)
        session.add(group)
        session.commit()


def get_groups_without_discipline(discipline_id: int) -> list[Group]:
    with db.Session() as session:
        smt = select(Group).where(
            ~Group.disciplines.any(id=discipline_id),
        )
        return session.scalars(smt).all()


def get_academic_performance(
    student_id: int, discipline_id: int
) -> tuple[str, int, int]:
    """
    Возвращает полное имя студента вместе с количеством пропущенных
    им занятий и общим количеством занятий для заданной дисциплины.
    """
    with db.Session() as session:
        student = session.get(Student, student_id)
        missed_classes_query = (
            session.query(
                func.sum(
                    MissedClass.is_missed.cast(Integer),
                ).label("missed_count"),
                func.count(MissedClass.id).label("total_count"),
            )
            .filter(
                and_(
                    MissedClass.student_id == student_id,
                    MissedClass.discipline_id == discipline_id,
                )
            )
            .group_by(MissedClass.student_id)
        )

        result = missed_classes_query.one_or_none()

        if result:
            missed_count, total_count = result
        else:
            missed_count, total_count = 0, 0

        return student.full_name, missed_count, total_count


def count_missed_classes(
    group_id: int, discipline_id: int
) -> tuple[int, list[AcademicPerformance]]:
    # используется для формирования краткого отчета
    with db.Session() as session:
        missed_classes_query = (
            session.query(
                Student.full_name.label("student_name"),
                func.sum(
                    MissedClass.is_missed.cast(Integer),
                ).label("missed_count"),
                func.count(MissedClass.id).label("total_count"),
            )
            .join(MissedClass, MissedClass.student_id == Student.id)
            .filter(Student.group_id == group_id)
            .filter(MissedClass.discipline_id == discipline_id)
            .group_by(Student.full_name)
            .order_by(Student.full_name)
        )

        results = missed_classes_query.all()

        missed_classes_by_group_and_discipline = [
            AcademicPerformance(
                student_name=student_name,
                number_of_passes=missed_count,
            )
            for student_name, missed_count, total_count in results
        ]

        return results[0][2], missed_classes_by_group_and_discipline


def missed_classes_with_day(
    group_id: int, discipline_id: int
) -> list[FullAcademicPerformance]:
    # используется для формирования полного отчета
    with db.Session() as session:
        missed_classes_query = (
            session.query(
                Student.full_name.label("student_name"),
                MissedClass.is_missed.label("missed_count"),
                MissedClass.date.label("date"),
            )
            .join(MissedClass, MissedClass.student_id == Student.id)
            .filter(
                and_(
                    Student.group_id == group_id,
                    MissedClass.discipline_id == discipline_id,
                )
            )
            .order_by(Student.full_name, MissedClass.date)
        )

        results = missed_classes_query.all()

        students: dict[str, list[MissedInfo]] = {}

        for student_name, is_missed, day in results:
            if student_name not in students:
                students[student_name] = []

            students[student_name].append(
                MissedInfo(
                    is_missed=is_missed,
                    day=day,
                )
            )

        result_list: list[FullAcademicPerformance] = [
            FullAcademicPerformance(
                student_name=name,
                missed_data=missed,
            )
            for name, missed in students.items()
        ]

        return result_list


def set_discipline2group(discipline_id: int, group_id: int) -> None:
    with db.Session() as session:
        discipline = session.get(Discipline, discipline_id)
        group = session.get(Group, group_id)
        group.disciplines.append(discipline)
        session.commit()
#
def is_student_registered(student_id: int) -> bool:
    with db.Session() as session:
        student = session.get(Student, student_id)
        return student.is_registered


def set_student_registered(student_id: int) -> None:
    with db.Session() as session:
        student = session.get(Student, student_id)
        student.is_registered = True
        session.commit()


def link_telegram_account(student_id: int, telegram_id: int) -> None:
    with db.Session() as session:
        student = session.get(Student, student_id)
        student.telegram_id = telegram_id
        session.commit()

def get_student_by_name_and_group(full_name: str, group_id: int) -> Student:
    with db.Session() as session:
        smt = select(Student).where(
            and_(
                Student.full_name == full_name,
                Student.group_id == group_id
            )
        )
        return session.scalars(smt).one_or_none()

def get_group_by_name(group_name: str) -> Group:
    with db.Session() as session:
        smt = select(Group).where(Group.name == group_name)
        return session.scalars(smt).one_or_none()
    
def create_team(name: str, discipline_id: int, group_id: int, creator_id: int) -> Team:
    teams = load_teams_from_excel()
    new_id = max((team.id for team in teams), default=0) + 1  # Генерируем уникальный идентификатор
    team = Team(
        id=new_id,
        name=name,
        discipline_id=discipline_id,
        group_id=group_id,
        creator_id=creator_id,
        members="",  # Инициализация пустой строкой
        reports="",  # Инициализация пустой строкой
        status=1,  # Инициализация открыт
        student_comment="",  # Инициализация пустой строкой
        teacher_comment=""  # Инициализация пустой строкой
    )
    teams.append(team)
    save_teams_to_excel(teams)
    return team

def join_team(team_id: int, student_id: int):
    teams = load_teams_from_excel()
    for team in teams:
        if team.id == team_id:
            if team.members is None:
                team.members = ""
                if str(student_id) not in team.members.split(";"):
                    team.members += f";{student_id}"
    save_teams_to_excel(teams)


def get_team_by_name_and_discipline(name: str, discipline_id: int, group_id: int) -> Team:
    with db.Session() as session:
        smt = select(Team).where(and_(Team.name == name, Team.discipline_id == discipline_id, Team.group_id == group_id))
        return session.scalars(smt).one_or_none()

def get_open_teams_by_discipline_and_group(discipline_id: int, group_id: int) -> list:
    teams = load_teams_from_excel()
    return [team for team in teams if team.discipline_id == discipline_id and team.group_id == group_id and team.status == 1]

def get_student_by_telegram_id(telegram_id: int) -> Student:
    with db.Session() as session:
        smt = select(Student).where(Student.telegram_id == telegram_id)
        return session.scalars(smt).one_or_none()

def close_team(team_id: int) -> None:
    with db.Session() as session:
        team = session.get(Team, team_id)
        team.close_team()
        session.commit()

def get_team_members(team_id: int) -> list[Student]:
    with db.Session() as session:
        team = session.get(Team, team_id)
        return team.students

def get_next_team_number(discipline_id: int, group_id: int) -> int:
    with db.Session() as session:
        smt = select(func.count(Team.id)).where(and_(Team.discipline_id == discipline_id, Team.group_id == group_id))
        return session.scalar(smt) + 1

def get_teams_by_student_id(student_id: int) -> list[Team]:
    with db.Session() as session:
        smt = select(Team).join(Team.students).where(Student.id == student_id)
        return session.scalars(smt).all()

def get_team_by_student_and_discipline(student_id: int, discipline_id: int) -> Team:
    with db.Session() as session:
        smt = select(Team).join(Team.students).where(Student.id == student_id, Team.discipline_id == discipline_id)
        return session.scalar(smt)

def get_discipline_by_id(discipline_id: int) -> Discipline:
    with db.Session() as session:
        return session.get(Discipline, discipline_id)

def save_teams_to_excel(teams: List[Team]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['ID', 'Name', 'Discipline_ID', 'Group_ID', 'Creator_ID', 'Members', 'Reports', 'Status', 'Student_Comment', 'Teacher_Comment'])

    for team in teams:
        sheet.append([team.id, team.name, team.discipline_id, team.group_id, team.creator_id, team.members, team.reports, team.status, team.student_comment, team.teacher_comment])

    workbook.save(PATH_TO_TEAMS)

def load_teams_from_excel() -> List[Team]:
    if not os.path.exists(PATH_TO_TEAMS):
        raise FileNotFoundError(f"Файл с командами не найден: {PATH_TO_TEAMS}")

    workbook = openpyxl.load_workbook(PATH_TO_TEAMS)
    sheet = workbook.active
    teams = []

    for row in sheet.iter_rows(min_row=2):
        team = Team(
            id=row[0].value,
            name=row[1].value,
            discipline_id=row[2].value,
            group_id=row[3].value,
            creator_id=row[4].value,
            members=row[5].value if row[5].value is not None else "",
            reports=row[6].value if len(row) > 6 and row[6].value is not None else "",
            status=row[7].value if len(row) > 7 and row[7].value is not None else 1,
            student_comment=row[8].value if len(row) > 8 and row[8].value is not None else "",
            teacher_comment=row[9].value if len(row) > 9 and row[9].value is not None else ""
        )
        teams.append(team)

    return teams

def get_teams_by_student_id(student_id: int) -> list:
    teams = load_teams_from_excel()
    student_teams = []

    for team in teams:
        if any(student.id == student_id for student in team.students):
            student_teams.append(team)

    return student_teams

def get_team_by_id(team_id: int) -> Team:
    teams = load_teams_from_excel()
    for team in teams:
        if team.id == team_id:
            return team
    return None

def user_has_team_in_discipline(student_id: int, discipline_id: int) -> bool:
    teams = load_teams_from_excel()
    for team in teams:
        if team.creator_id == student_id and team.discipline_id == discipline_id:
            return True
    return False

def get_teams_created_by_user(student_id: int) -> list:
    teams = load_teams_from_excel()
    user_teams = [team for team in teams if team.creator_id == student_id]
    return user_teams

def save_report(team_id: int, report: str):
    teams = load_teams_from_excel()
    for team in teams:
        if team.id == team_id:
            team.reports = report
            break
    save_teams_to_excel(teams)

def get_students_by_ids(ids: list) -> list:
    with db.Session as session:
        return session.query(Student).filter(Student.id.in_(ids)).all()
    
def update_team_comment(team_id: int, comment: str, comment_type: str):
    teams = load_teams_from_excel()
    for team in teams:
        if team.id == team_id:
            if comment_type == "student":
                team.student_comment = comment
            elif comment_type == "teacher":
                team.teacher_comment = comment
            break
    save_teams_to_excel(teams)

    
def get_teams_by_discipline(discipline_id: int) -> list:
    teams = load_teams_from_excel()
    result = [team for team in teams if team.discipline_id == discipline_id]

    print(f"Loaded Teams: {teams}")
    print(f"Filtered Teams for Discipline ID {discipline_id}: {result}")
    
    if not result:
        print(f"No teams found for discipline_id: {discipline_id}")
    
    return result

def get_unique_disciplines() -> List[Dict[str, str]]:
    workbook = openpyxl.load_workbook(PATH_TO_CONFIG_DATA)
    disciplines = set()
    
    for sheet_name in workbook.sheetnames:
        group, discipline = sheet_name.split('|')
        disciplines.add(discipline.strip())
    
    return [{'id': idx, 'name': discipline} for idx, discipline in enumerate(disciplines, start=1)]

def get_groups_by_discipline(discipline_name: str) -> List[Dict[str, str]]:
    workbook = openpyxl.load_workbook(PATH_TO_CONFIG_DATA)
    groups = set()
    
    for sheet_name in workbook.sheetnames:
        group, discipline = sheet_name.split('|')
        if discipline.strip() == discipline_name:
            groups.add(group.strip())
    
    return [{'id': idx, 'name': group} for idx, group in enumerate(groups, start=1)]

def get_teams_by_group(group_name: str, discipline_name: str) -> List[Team]:
    # Словарь для отображения имени группы на ID группы
    group_id_map = {
        '4316': 1,
        '4317': 2
        # Добавьте сюда другие группы по мере необходимости
    }
    
    # Словарь для отображения имени дисциплины на ID дисциплины
    discipline_id_map = {
        'ОПД': 1,
        'АЛГ': 2
        # Добавьте сюда другие дисциплины по мере необходимости
    }
    
    group_id = group_id_map.get(group_name)
    discipline_id = discipline_id_map.get(discipline_name)
    
    teams = load_teams_from_excel()
    filtered_teams = [team for team in teams if team.group_id == group_id and team.discipline_id == discipline_id]
    print(f"Loaded teams: {teams}")  # Отладочная информация
    print(f"Filtered teams for group {group_name} (ID {group_id}) and discipline {discipline_name} (ID {discipline_id}): {filtered_teams}")  # Отладочная информация
    return filtered_teams




